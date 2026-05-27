"""Phase 1C: add 'Newsroom RSS' column to URL Registry and populate confirmed feeds.

Per the Phase 1B discovery sweep on the 4 test names:
  - WIX: live feed (latest entry May 13 2026) → populate
  - SHOP: feed exists but stale (last entry Dec 2023) → leave blank, use HTML
  - DASH, CART: no RSS → leave blank, use HTML

Other 31 names: column added but blank — will be filled as RSS is verified
per name on demand.
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "config" / "company_urls_jc.xlsx"

# Confirmed live feeds (verified to return current entries; not stale)
CONFIRMED_FEEDS: dict[str, str] = {
    "WIX": "https://www.wix.com/press-room/home/blog-feed.xml",
}


def main() -> int:
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["URL Registry"]
    header = [c.value for c in ws[1]]

    col_name = "Newsroom RSS"
    if col_name in header:
        print(f"Column '{col_name}' already exists. Updating values only.")
        col = header.index(col_name) + 1
    else:
        col = ws.max_column + 1
        cell = ws.cell(row=1, column=col, value=col_name)
        cell.font = Font(bold=True)
        print(f"Added column '{col_name}' at position {col}")

    ticker_col = header.index("Ticker") + 1 if "Ticker" in header else 1
    populated = 0
    for row_idx in range(2, ws.max_row + 1):
        tkr = ws.cell(row=row_idx, column=ticker_col).value
        if not tkr:
            continue
        feed = CONFIRMED_FEEDS.get(str(tkr).strip().upper())
        if feed:
            ws.cell(row=row_idx, column=col, value=feed)
            populated += 1

    wb.save(XLSX)
    print(f"Populated {populated} confirmed feed URLs: {list(CONFIRMED_FEEDS)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
