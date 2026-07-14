"""Mechanical URL audit for v4 failures.

For each known-broken (ticker, source) pair, try a set of common URL
patterns against the company's domain and pick the first one that
returns 200 + real content. Bot-blocked names (TSLA, UBER, PAYC) get
their URLs blanked entirely so they stop trying and log clean skips.

Runs in ~2-3 minutes total (~150 HEAD requests with 0.15s sleep).
"""

from __future__ import annotations

import sys
import time
from pathlib import Path

import openpyxl
import requests
from openpyxl.worksheet.hyperlink import Hyperlink

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "config" / "company_urls_jc.xlsx"
UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_0) AppleWebKit/537.36 consumer-monitor/0.1"
TIMEOUT = 15

# Per-source URL patterns to try, ordered by likelihood
PATTERNS = {
    "management_team": [
        "/about/leadership", "/company/leadership", "/leadership",
        "/team", "/about-us/leadership", "/about/team",
        "/governance/leadership", "/corporate-governance/leadership",
        "/investors/governance/leadership", "/investor-relations/leadership",
    ],
    "newsroom": [
        "/news", "/newsroom", "/press", "/press-releases",
        "/company/news", "/media", "/about/news", "/media-centre",
        "/press-room", "/company/press", "/blog",
    ],
    "careers": [
        "/careers", "/jobs", "/company/careers", "/about/careers",
    ],
    "terms": [
        "/terms", "/legal", "/legal/terms", "/terms-of-use",
        "/terms-of-service", "/legal/terms-of-service",
        "/terms-and-conditions", "/legal/terms-of-use",
    ],
    "investor_relations": [
        "/investors", "/investor-relations", "/investor",
        "/company/investors", "/about/investors",
    ],
    "pricing": [
        "/pricing", "/plans", "/products",
    ],
}

# Failures to try to fix. Each entry: (ticker, source, list of candidate URLs to try).
# Candidates use the company's known root domain.
BROKEN = {
    "ABNB": {"management_team": "https://investors.airbnb.com"},
    "ADP":  {"terms": "https://www.adp.com"},
    "BBY":  {"newsroom": "https://corporate.bestbuy.com"},
    "BKNG": {
        "investor_relations": "https://www.bookingholdings.com",
        "newsroom": "https://www.bookingholdings.com",
        "careers": "https://www.bookingholdings.com",  # DNS-failed subdomain
    },
    "EXPE": {
        "management_team": "https://www.expediagroup.com",
        "newsroom": "https://www.expediagroup.com",  # DNS-failed subdomain
    },
    "FOXA": {"newsroom": "https://www.foxcorporation.com"},
    "GOOG": {"management_team": "https://abc.xyz"},
    "LAMR": {
        "investor_relations": "https://www.lamar.com",
        "newsroom": "https://www.lamar.com",
        "careers": "https://www.lamar.com",
        "terms": "https://www.lamar.com",
    },
    "LOGI": {
        "management_team": "https://ir.logitech.com",
        "newsroom": "https://www.logitech.com",
        "terms": "https://www.logitech.com",
    },
    "LYFT": {
        "management_team": "https://investor.lyft.com",
        "pricing": "https://www.lyft.com",
    },
    "LYV": {
        "terms": "https://www.livenationentertainment.com",
        "careers": "https://www.livenationentertainment.com",  # DNS-failed subdomain
    },
    "NFLX": {"newsroom": "https://about.netflix.com"},
    "OUT":  {
        "newsroom": "https://www.outfrontmedia.com",
        "terms": "https://www.outfrontmedia.com",
    },
    "PAYX": {"terms": "https://www.paychex.com"},
    "RBLX": {"newsroom": "https://ir.roblox.com"},
    "SNAP": {"management_team": "https://investor.snap.com"},
    "SPHR": {
        "investor_relations": "https://www.sphereentertainmentco.com",
        "management_team": "https://www.sphereentertainmentco.com",
        "newsroom": "https://www.sphereentertainmentco.com",
        "terms": "https://www.sphereentertainmentco.com",
    },
    "TTWO": {"newsroom": "https://ir.take2games.com"},
    "U":    {"newsroom": "https://unity.com"},
    "UMG":  {"investor_relations": "https://www.universalmusic.com"},
    "WMG":  {"careers": "https://www.wmg.com"},  # DNS-failed subdomain
    "YOU":  {"terms": "https://www.clearme.com"},
    # ABNB investor_relations was suspect-empty (bot detection) not 404 — skip
}

# Names whose URLs are permanently bot-blocked from GH runners. Clear these
# entirely so the scraper stops trying and logs clean skips.
CLEAR_URLS = {
    "TSLA": ["terms", "management_team", "pricing", "newsroom",
             "investor_relations", "careers"],
    "UBER": ["terms", "pricing", "newsroom", "careers"],
    "PAYC": ["terms", "newsroom", "careers"],
}


def _head(url: str) -> tuple[int, int]:
    """Return (status_code, content_length_estimate). content_length -1 if HEAD."""
    try:
        r = requests.get(url, headers={"User-Agent": UA}, timeout=TIMEOUT,
                         allow_redirects=True, stream=True)
        # Read a small chunk to see if there's real content vs bot-stub
        chunk = next(r.iter_content(4096), b"")
        r.close()
        return r.status_code, len(chunk)
    except requests.RequestException:
        return 0, 0


def find_working(base_url: str, source: str) -> str | None:
    patterns = PATTERNS.get(source, [])
    for p in patterns:
        candidate = base_url.rstrip("/") + p
        status, size = _head(candidate)
        if status == 200 and size > 500:
            print(f"    ✓ {candidate}  ({size} bytes)")
            return candidate
        time.sleep(0.15)
    return None


def main() -> int:
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["URL Registry"]
    header = [c.value for c in ws[1]]
    col = {h: i + 1 for i, h in enumerate(header) if h}
    ticker_col = col["Ticker"]
    source_col = {
        "management_team": col["Management Team"],
        "newsroom": col["Newsroom / Press"],
        "careers": col["Careers / Jobs"],
        "terms": col["Terms & Conditions"],
        "investor_relations": col["Investor Relations"],
        "pricing": col["Product / Pricing"],
    }
    ticker_rows = {}
    for r in range(2, ws.max_row + 1):
        t = str(ws.cell(row=r, column=ticker_col).value or "").strip()
        if t:
            ticker_rows[t] = r

    fixed = 0
    skipped = 0
    print("=== Auditing broken URLs ===\n")
    for ticker, sources in BROKEN.items():
        if ticker not in ticker_rows:
            continue
        r = ticker_rows[ticker]
        for source, base in sources.items():
            print(f"  {ticker}/{source}  base={base}")
            found = find_working(base, source)
            if found:
                cell = ws.cell(row=r, column=source_col[source])
                cell.value = found
                if cell.hyperlink is not None:
                    cell.hyperlink.target = found
                else:
                    cell.hyperlink = Hyperlink(ref=cell.coordinate, target=found)
                fixed += 1
            else:
                print(f"    ✗ no working candidate found")
                skipped += 1

    print(f"\nFixed: {fixed}  Skipped: {skipped}")

    print("\n=== Clearing bot-blocked URLs ===")
    cleared = 0
    for ticker, sources in CLEAR_URLS.items():
        if ticker not in ticker_rows:
            continue
        r = ticker_rows[ticker]
        for source in sources:
            cell = ws.cell(row=r, column=source_col[source])
            if cell.value:
                cell.value = None
                cell.hyperlink = None
                cleared += 1
    print(f"Cleared {cleared} bot-blocked URLs from TSLA/UBER/PAYC")

    wb.save(XLSX)
    print("\nSaved.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
