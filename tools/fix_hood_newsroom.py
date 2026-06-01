"""One-shot xlsx fix: HOOD newsroom URL points at the bot-blocked
newsroom.robinhood.com. Replace with newsroom.aboutrobinhood.com which
returns actual press releases (verified 2026-06-01 — 2067 chars extracted)."""

from pathlib import Path
import openpyxl
from openpyxl.worksheet.hyperlink import Hyperlink

XLSX = Path(__file__).resolve().parent.parent / "config" / "company_urls_jc.xlsx"
NEW_URL = "https://newsroom.aboutrobinhood.com"

wb = openpyxl.load_workbook(XLSX)
ws = wb["URL Registry"]
header = [c.value for c in ws[1]]
ticker_col = header.index("Ticker") + 1
newsroom_col = header.index("Newsroom / Press") + 1

for row in range(2, ws.max_row + 1):
    if str(ws.cell(row=row, column=ticker_col).value or "").strip() == "HOOD":
        cell = ws.cell(row=row, column=newsroom_col)
        old = cell.value
        cell.value = NEW_URL
        if cell.hyperlink is not None:
            cell.hyperlink.target = NEW_URL
        else:
            cell.hyperlink = Hyperlink(ref=cell.coordinate, target=NEW_URL)
        print(f"HOOD newsroom: {old}")
        print(f"            -> {NEW_URL}")
        break

wb.save(XLSX)
print("Saved.")
