"""Mechanical URL fix for the medium-tier fixable failures.

Only targets the IP-independent failures from the 24-name health check —
dead/unreachable hosts (connection errors) and 404s — where a correct URL
almost certainly exists on the company's own site. Bot-blocked (HTTP 403)
and JS-rendered (suspect-empty) sources are intentionally excluded: those
can't be fixed by swapping a URL.

For each (ticker, source) it probes common URL patterns against the
company's real root domain and keeps the first that returns 200 + real
content. A few sources that genuinely don't exist (e.g. a static pricing
page for a marketplace) are blanked so they stop erroring.

Mirrors tools/audit_and_fix_v4.py.
"""

from __future__ import annotations

import sys
import time
from pathlib import Path

import openpyxl
import requests
from openpyxl.worksheet.hyperlink import Hyperlink

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "config" / "company_urls_jc.xlsx"
UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_0) AppleWebKit/537.36 consumer-monitor/0.1"
TIMEOUT = 15

PATTERNS = {
    "management_team": [
        "/governance/leadership", "/corporate-governance/leadership",
        "/leadership", "/company/leadership", "/about/leadership",
        "/team", "/about-us/leadership", "/about/team",
        "/investors/governance/leadership", "/our-team", "/about/management",
    ],
    "newsroom": [
        "/news", "/newsroom", "/press", "/press-releases",
        "/news-events/press-releases", "/company/news", "/media",
        "/about/news", "/press-room", "/company/press", "/blog", "/news-events",
    ],
    "careers": [
        "/careers", "/jobs", "/company/careers", "/about/careers", "/careers/",
    ],
    "terms": [
        "/terms", "/legal", "/legal/terms", "/terms-of-use",
        "/terms-of-service", "/legal/terms-of-service",
        "/terms-and-conditions", "/legal/terms-of-use", "/policies/terms",
    ],
    "investor_relations": [
        "", "/investors", "/investor-relations", "/investor",
        "/overview/default.aspx", "/home/default.aspx",
    ],
    "pricing": ["/pricing", "/plans", "/products"],
}

# Fixable failures only (connection errors + 404s). base = company root domain.
BROKEN = {
    "AFRM": {"investor_relations": "https://investors.affirm.com",
             "management_team": "https://investors.affirm.com",
             "newsroom": "https://www.affirm.com"},
    "CART": {"investor_relations": "https://investors.instacart.com",
             "management_team": "https://investors.instacart.com"},
    "DAVE": {"investor_relations": "https://investors.dave.com",
             "management_team": "https://investors.dave.com"},
    "UPST": {"investor_relations": "https://ir.upstart.com",
             "management_team": "https://ir.upstart.com",
             "terms": "https://www.upstart.com"},
    "GLBE": {"investor_relations": "https://investors.global-e.com",
             "management_team": "https://investors.global-e.com",
             "careers": "https://www.global-e.com"},
    "WSM":  {"investor_relations": "https://ir.williams-sonomainc.com",
             "management_team": "https://ir.williams-sonomainc.com",
             "newsroom": "https://ir.williams-sonomainc.com",
             "careers": "https://www.williams-sonoma.com"},
    "SEZL": {"management_team": "https://sezzle.com",
             "newsroom": "https://sezzle.com"},
    "ENVA": {"careers": "https://www.enova.com",
             "newsroom": "https://www.enova.com"},
    "SGI":  {"newsroom": "https://www.somnigroup.com",
             "terms": "https://www.somnigroup.com"},
    "W":    {"newsroom": "https://investor.wayfair.com",
             "terms": "https://www.wayfair.com"},
    "MTCH": {"newsroom": "https://mtch.com",
             "careers": "https://mtch.com"},
    "ENVA_dup": {},  # placeholder guard (unused)
}

# Sources that genuinely have no such static page → blank them so they stop erroring.
CLEAR_URLS = {
    "DASH": ["pricing"],   # DoorDash has no static consumer pricing page
    "W": ["pricing"],      # Wayfair — no single pricing page
    "SGI": ["pricing"],    # Somnigroup — no pricing page
}


def _get(url: str) -> tuple[int, int]:
    try:
        r = requests.get(url, headers={"User-Agent": UA}, timeout=TIMEOUT,
                         allow_redirects=True, stream=True)
        chunk = next(r.iter_content(4096), b"")
        r.close()
        return r.status_code, len(chunk)
    except requests.RequestException:
        return 0, 0


def find_working(base_url: str, source: str) -> str | None:
    for p in PATTERNS.get(source, []):
        candidate = base_url.rstrip("/") + p
        status, size = _get(candidate)
        if status == 200 and size > 500:
            print(f"    ✓ {candidate}  ({size} bytes)")
            return candidate
        time.sleep(0.15)
    return None


def main() -> int:
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["URL Registry"]
    header = [c.value for c in ws[1]]
    col = {h: i + 1 for i, h in enumerate(header) if h}
    ticker_col = col["Ticker"]
    source_col = {
        "management_team": col["Management Team"],
        "newsroom": col["Newsroom / Press"],
        "careers": col["Careers / Jobs"],
        "terms": col["Terms & Conditions"],
        "investor_relations": col["Investor Relations"],
        "pricing": col["Product / Pricing"],
    }
    ticker_rows = {}
    for r in range(2, ws.max_row + 1):
        t = str(ws.cell(row=r, column=ticker_col).value or "").strip()
        if t:
            ticker_rows[t] = r

    fixed = skipped = 0
    print("=== Probing fixable medium-tier URLs ===\n")
    for ticker, sources in BROKEN.items():
        if ticker not in ticker_rows or not sources:
            continue
        r = ticker_rows[ticker]
        for source, base in sources.items():
            print(f"  {ticker}/{source}  base={base}")
            found = find_working(base, source)
            if found:
                cell = ws.cell(row=r, column=source_col[source])
                cell.value = found
                if cell.hyperlink is not None:
                    cell.hyperlink.target = found
                else:
                    cell.hyperlink = Hyperlink(ref=cell.coordinate, target=found)
                fixed += 1
            else:
                print("    ✗ no working candidate")
                skipped += 1

    print(f"\nFixed: {fixed}  Skipped: {skipped}")

    print("\n=== Blanking non-existent pages ===")
    cleared = 0
    for ticker, sources in CLEAR_URLS.items():
        if ticker not in ticker_rows:
            continue
        r = ticker_rows[ticker]
        for source in sources:
            cell = ws.cell(row=r, column=source_col[source])
            if cell.value:
                cell.value = None
                cell.hyperlink = None
                cleared += 1
                print(f"  cleared {ticker}/{source}")
    print(f"Cleared {cleared}")

    wb.save(XLSX)
    print("\nSaved.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
