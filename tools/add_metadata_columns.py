"""Phase 1A: add Priority Tier + Active columns to the URL Registry tab.

Minimal additive change — existing data untouched. Tier values sourced from
scrapers/tiers.py for backward compatibility; Active defaults to TRUE.
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "config" / "company_urls_jc.xlsx"

sys.path.insert(0, str(ROOT))
from scrapers.tiers import tier_for  # noqa: E402

NEW_COLUMNS = ["Priority Tier", "Active"]


def main() -> int:
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["URL Registry"]

    header = [c.value for c in ws[1]]

    # Check if columns already exist
    already = [c for c in NEW_COLUMNS if c in header]
    if already:
        print(f"Columns already exist: {already}. Aborting.", file=sys.stderr)
        return 1

    # Append new column headers
    start_col = ws.max_column + 1
    bold = Font(bold=True)
    for i, name in enumerate(NEW_COLUMNS):
        cell = ws.cell(row=1, column=start_col + i, value=name)
        cell.font = bold

    # Build header index after adding
    header = [c.value for c in ws[1]]
    col_idx = {h: i + 1 for i, h in enumerate(header) if h}
    tier_col = col_idx["Priority Tier"]
    active_col = col_idx["Active"]
    ticker_col = col_idx["Ticker"]

    # Populate per row
    populated = 0
    for row_idx in range(2, ws.max_row + 1):
        tkr = ws.cell(row=row_idx, column=ticker_col).value
        if not tkr:
            continue
        ws.cell(row=row_idx, column=tier_col, value=tier_for(str(tkr)))
        ws.cell(row=row_idx, column=active_col, value=True)
        populated += 1

    wb.save(XLSX)
    print(f"Added {len(NEW_COLUMNS)} columns: {NEW_COLUMNS}")
    print(f"Populated {populated} rows")
    return 0


if __name__ == "__main__":
    sys.exit(main())
