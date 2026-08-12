"""Encode the two-analyst coverage split as the source of truth in the xlsx.

Adds/updates an "Analyst" column: "1" or "2" per name. Names in neither
analyst's list are set Active=FALSE (kept as rows so they can be re-added).
Verifies every listed ticker exists before writing; aborts otherwise.
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

XLSX = Path(__file__).resolve().parent.parent / "config" / "company_urls_jc.xlsx"

ANALYST_1 = [
    "BKNG", "EXPE", "ABNB", "YOU", "CHH", "META", "GOOG", "SNAP", "PINS", "RDDT",
    "APP", "U", "UBER", "LYFT", "TSLA", "NFLX", "SPOT", "NYT", "LIF", "LOGI",
    "GRMN", "BBY", "DIS", "FOXA", "UMG", "WMG", "LYV", "SPHR", "STUB", "LION",
    "OUT", "LAMR", "TTWO", "EA", "RBLX", "INTU", "ADP", "PAYX", "PAYC", "PCTY", "CDW",
]
ANALYST_2 = [
    "SHOP", "WIX", "GLBE", "GDDY", "EBAY", "W", "ETSY", "CHWY", "DASH", "CART",
    "DUOL", "MTCH", "Z", "CSGP", "RKT", "COMP", "APPF", "XYZ", "BILL", "PYPL",
    "TOST", "ADYEN NA", "SEZL", "DAVE", "ZIP AU", "AFRM", "KLAR", "HOOD", "UPST",
    "RH", "WSM", "SGI", "ENVA", "ONON", "8136 JP",
]


def main() -> int:
    assigned = {t: "1" for t in ANALYST_1}
    for t in ANALYST_2:
        if t in assigned:
            print(f"ERROR: {t} is in BOTH analyst lists", file=sys.stderr)
            return 1
        assigned[t] = "2"

    wb = openpyxl.load_workbook(XLSX)
    ws = wb["URL Registry"]
    header = [c.value for c in ws[1]]
    col = {h: i + 1 for i, h in enumerate(header) if h}
    ticker_col = col["Ticker"]
    active_col = col["Active"]

    # Ensure an "Analyst" column exists (append if missing).
    if "Analyst" not in col:
        analyst_col = ws.max_column + 1
        ws.cell(row=1, column=analyst_col, value="Analyst")
    else:
        analyst_col = col["Analyst"]

    xlsx_tickers = {}
    for r in range(2, ws.max_row + 1):
        t = str(ws.cell(row=r, column=ticker_col).value or "").strip()
        if t:
            xlsx_tickers[t] = r

    missing = [t for t in assigned if t not in xlsx_tickers]
    if missing:
        print(f"ERROR: listed tickers not found in xlsx: {missing}", file=sys.stderr)
        return 1

    a1 = a2 = deactivated = 0
    leftover = []
    for t, r in xlsx_tickers.items():
        who = assigned.get(t)
        ws.cell(row=r, column=analyst_col, value=who)  # None if unassigned
        if who == "1":
            a1 += 1
        elif who == "2":
            a2 += 1
        else:
            # In neither analyst's coverage → drop from the active universe.
            ws.cell(row=r, column=active_col, value=False)
            deactivated += 1
            leftover.append(t)

    wb.save(XLSX)
    print(f"Analyst 1: {a1} names")
    print(f"Analyst 2: {a2} names")
    print(f"Deactivated (in neither list): {deactivated} -> {sorted(leftover)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
