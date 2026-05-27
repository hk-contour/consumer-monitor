"""Phase 1D: add per-source CSS selector columns to URL Registry.

Selectors narrow what static_pages.fetch_and_extract pulls from each page,
cutting noise from sidebars, recommendation widgets, personalization, etc.

Starting scope: pricing pages for the 4 test names. Other source types and
other names get blank cells and use the default whole-body extraction.
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "config" / "company_urls_jc.xlsx"

# One new column per source type that supports selectors. For now: pricing only.
# Column header → {ticker: selector}
SELECTOR_COLUMNS: dict[str, dict[str, str]] = {
    "Pricing Selector": {
        "SHOP": "main",
        "WIX":  "main",
        "CART": "main",
        # DASH skipped: pricing URL currently 404s
    },
}


def main() -> int:
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["URL Registry"]
    header = [c.value for c in ws[1]]
    col_idx = {h: i + 1 for i, h in enumerate(header) if h}
    ticker_col = col_idx.get("Ticker", 1)

    for col_name, values in SELECTOR_COLUMNS.items():
        if col_name in col_idx:
            print(f"Column '{col_name}' exists. Updating values.")
            col = col_idx[col_name]
        else:
            col = ws.max_column + 1
            cell = ws.cell(row=1, column=col, value=col_name)
            cell.font = Font(bold=True)
            col_idx[col_name] = col
            print(f"Added column '{col_name}' at position {col}")

        populated = 0
        for row_idx in range(2, ws.max_row + 1):
            tkr = str(ws.cell(row=row_idx, column=ticker_col).value or "").strip()
            if tkr in values:
                ws.cell(row=row_idx, column=col, value=values[tkr])
                populated += 1
        print(f"  Populated {populated} cells with selectors")

    wb.save(XLSX)
    return 0


if __name__ == "__main__":
    sys.exit(main())
