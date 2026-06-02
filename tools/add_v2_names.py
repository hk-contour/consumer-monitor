"""Add the v2 coverage set (consumer hardware / travel / ride-share / experiential)
to the URL Registry tab. Tier = 'v2' so they don't enter --tier high/medium/all
filters; use --ticker to scrape them explicitly.

URL guesses use common patterns (investor.{co}.com, etc.) — some will 404 on
first run and we'll fix in a follow-up pass. EDGAR works regardless because
config_reader uses SEC's ticker→CIK index.
"""

from pathlib import Path
import openpyxl
from openpyxl.worksheet.hyperlink import Hyperlink

XLSX = Path(__file__).resolve().parent.parent / "config" / "company_urls_jc.xlsx"

# Each row: ticker, company, exchange, IR, newsroom, mgmt, careers, pricing,
# terms, reg filings (SEC), reddit, notes, tier, active
ROWS = [
    {
        "ticker": "LOGI", "company": "Logitech International S.A.", "exchange": "NASDAQ",
        "ir": "https://ir.logitech.com",
        "newsroom": "https://www.logitech.com/en-us/about/newsroom.html",
        "management_team": "https://ir.logitech.com/corporate-governance/leadership-team",
        "careers": "https://jobs.logitech.com",
        "pricing": "",
        "terms": "https://www.logitech.com/en-us/legal/terms-of-use.html",
        "reg": "https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK=LOGI",
        "reddit": "",
        "notes": "Consumer hardware (mice, keyboards, webcams); watch G PRO gaming and Streamlabs ecosystem",
    },
    {
        "ticker": "GRMN", "company": "Garmin Ltd.", "exchange": "NYSE",
        "ir": "https://www.garmin.com/en-US/company/investors/",
        "newsroom": "https://newsroom.garmin.com",
        "management_team": "https://www.garmin.com/en-US/company/leadership/",
        "careers": "https://careers.garmin.com",
        "pricing": "",
        "terms": "https://www.garmin.com/en-US/legal/terms-of-use/",
        "reg": "https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK=GRMN",
        "reddit": "https://www.reddit.com/r/Garmin/",
        "notes": "GPS, wearables, marine, aviation; watch fitness wearable segment vs Apple Watch",
    },
    {
        "ticker": "TSLA", "company": "Tesla, Inc.", "exchange": "NASDAQ",
        "ir": "https://ir.tesla.com",
        "newsroom": "https://www.tesla.com/blog",
        "management_team": "https://ir.tesla.com/corporate/leadership",
        "careers": "https://www.tesla.com/careers",
        "pricing": "https://www.tesla.com/inventory/new/m3",
        "terms": "https://www.tesla.com/about/legal",
        "reg": "https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK=TSLA",
        "reddit": "https://www.reddit.com/r/teslamotors/",
        "notes": "Watch FSD pricing, vehicle MSRP changes, energy storage deployments, Cybertruck/Semi shipments",
    },
    {
        "ticker": "LYFT", "company": "Lyft, Inc.", "exchange": "NASDAQ",
        "ir": "https://investor.lyft.com",
        "newsroom": "https://www.lyft.com/blog",
        "management_team": "https://investor.lyft.com/corporate-governance/leadership",
        "careers": "https://www.lyft.com/careers",
        "pricing": "https://www.lyft.com/pricing",
        "terms": "https://www.lyft.com/terms",
        "reg": "https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK=LYFT",
        "reddit": "https://www.reddit.com/r/lyftdrivers/",
        "notes": "Watch Lyft Pink subscription, driver pay, ride pricing; r/lyftdrivers leads on policy changes",
    },
    {
        "ticker": "UBER", "company": "Uber Technologies, Inc.", "exchange": "NYSE",
        "ir": "https://investor.uber.com",
        "newsroom": "https://www.uber.com/newsroom/",
        "management_team": "https://investor.uber.com/governance/leadership",
        "careers": "https://www.uber.com/global/en/careers/",
        "pricing": "https://www.uber.com/us/en/ride/uberone/",
        "terms": "https://www.uber.com/legal/terms/",
        "reg": "https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK=UBER",
        "reddit": "https://www.reddit.com/r/uberdrivers/",
        "notes": "Watch Uber One subscription, freight segment, autonomous vehicle partnerships, delivery commission",
    },
    {
        "ticker": "EXPE", "company": "Expedia Group, Inc.", "exchange": "NASDAQ",
        "ir": "https://www.expediagroup.com/investors/",
        "newsroom": "https://newsroom.expediagroup.com",
        "management_team": "https://www.expediagroup.com/about/leadership/default.aspx",
        "careers": "https://careers.expediagroup.com",
        "pricing": "",
        "terms": "https://www.expedia.com/legal",
        "reg": "https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK=EXPE",
        "reddit": "https://www.reddit.com/r/expedia/",
        "notes": "Owns Vrbo, Hotels.com; watch ADR, room nights, take-rate changes, AI search rollout",
    },
    {
        "ticker": "BKNG", "company": "Booking Holdings Inc.", "exchange": "NASDAQ",
        "ir": "https://www.bookingholdings.com/investors/",
        "newsroom": "https://www.bookingholdings.com/news/",
        "management_team": "https://www.bookingholdings.com/about/leadership/",
        "careers": "https://careers.bookingholdings.com",
        "pricing": "",
        "terms": "https://www.booking.com/content/terms.html",
        "reg": "https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK=BKNG",
        "reddit": "",
        "notes": "Booking.com, Priceline, Agoda, Kayak; watch room-night growth, alternative-accommodations mix, Genius loyalty tier expansion",
    },
    {
        "ticker": "ABNB", "company": "Airbnb, Inc.", "exchange": "NASDAQ",
        "ir": "https://investors.airbnb.com",
        "newsroom": "https://news.airbnb.com",
        "management_team": "https://investors.airbnb.com/leadership/",
        "careers": "https://careers.airbnb.com",
        "pricing": "",
        "terms": "https://www.airbnb.com/terms",
        "reg": "https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK=ABNB",
        "reddit": "https://www.reddit.com/r/airbnb_hosts/",
        "notes": "Watch host fee changes, Experiences relaunch, regulatory pressure (EU/NYC), Co-Host marketplace",
    },
    {
        "ticker": "SPHR", "company": "Sphere Entertainment Co.", "exchange": "NYSE",
        "ir": "https://www.sphereentertainmentco.com/investors/",
        "newsroom": "https://www.sphereentertainmentco.com/newsroom/",
        "management_team": "https://www.sphereentertainmentco.com/governance/leadership/",
        "careers": "https://www.sphereentertainmentco.com/careers/",
        "pricing": "https://www.thesphere.com/upcoming-events",
        "terms": "https://www.sphereentertainmentco.com/legal/",
        "reg": "https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK=SPHR",
        "reddit": "",
        "notes": "Sphere Las Vegas venue + MSG Networks; watch event bookings, sponsor deals, Abu Dhabi sphere progress",
    },
]


def main() -> int:
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["URL Registry"]
    header = [c.value for c in ws[1]]
    col = {h: i + 1 for i, h in enumerate(header) if h}

    # Map xlsx header -> dict key
    write_map = {
        "Ticker": "ticker", "Company": "company", "Exchange": "exchange",
        "Investor Relations": "ir", "Newsroom / Press": "newsroom",
        "Management Team": "management_team", "Careers / Jobs": "careers",
        "Product / Pricing": "pricing", "Terms & Conditions": "terms",
        "Regulatory Filings": "reg", "Reddit Community": "reddit",
        "Monitoring Notes": "notes",
    }

    existing = {
        str(ws.cell(row=r, column=col["Ticker"]).value or "").strip()
        for r in range(2, ws.max_row + 1)
    }

    added = 0
    for row in ROWS:
        if row["ticker"] in existing:
            print(f"  {row['ticker']:6s} already in xlsx — skip")
            continue
        new_row = ws.max_row + 1
        for hdr, key in write_map.items():
            if hdr in col:
                val = row.get(key, "")
                cell = ws.cell(row=new_row, column=col[hdr], value=val)
                if val and val.startswith("http"):
                    cell.hyperlink = Hyperlink(ref=cell.coordinate, target=val)
        if "Priority Tier" in col:
            ws.cell(row=new_row, column=col["Priority Tier"], value="v2")
        if "Active" in col:
            ws.cell(row=new_row, column=col["Active"], value=True)
        print(f"  {row['ticker']:6s} added at row {new_row}")
        added += 1

    wb.save(XLSX)
    print(f"\nAdded {added} new names with tier='v2', Active=TRUE.")
    print("Trigger with: gh workflow run monitor.yml -f tickers=LOGI,GRMN,TSLA,LYFT,UBER,EXPE,BKNG,ABNB,SPHR")
    return 0


if __name__ == "__main__":
    import sys
    sys.exit(main())
