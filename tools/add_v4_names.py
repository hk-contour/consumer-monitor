"""Add the v4 coverage set — a broad consumer / media / travel / gaming /
HR-tech expansion. 32 new names beyond what's already in v1/v2/v3.

Priority Tier = 'v4'. Use --ticker to scrape.

URL guesses follow common patterns. Some 404 on first run; EDGAR
coverage works immediately for all US-listed names via SEC's ticker
index. UMG is Euronext Amsterdam (non-US, will have limited EDGAR).

LIF and LION are non-standard ticker labels — added with ticker as
placeholder; will likely fail EDGAR lookup, resolve later.
"""

from pathlib import Path
import openpyxl
from openpyxl.worksheet.hyperlink import Hyperlink

XLSX = Path(__file__).resolve().parent.parent / "config" / "company_urls_jc.xlsx"


def r(ticker, company, exchange="NASDAQ",
      ir="", newsroom="", management_team="", careers="",
      pricing="", terms="", reddit="", notes=""):
    reg = f"https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK={ticker}"
    return dict(ticker=ticker, company=company, exchange=exchange,
                ir=ir, newsroom=newsroom, management_team=management_team,
                careers=careers, pricing=pricing, terms=terms,
                reg=reg, reddit=reddit, notes=notes)


ROWS = [
    r("YOU",  "Clear Secure, Inc.", "NYSE",
      ir="https://ir.clearme.com", newsroom="https://ir.clearme.com/news-events",
      careers="https://www.clearme.com/careers", terms="https://www.clearme.com/legal",
      notes="Biometric identity verification; watch Clear Plus subscription growth, airport partnership expansions, TSA PreCheck competition"),
    r("CHH",  "Choice Hotels International, Inc.", "NYSE",
      ir="https://media.choicehotels.com/investors", newsroom="https://media.choicehotels.com",
      careers="https://careers.choicehotels.com", terms="https://www.choicehotels.com/legal/terms-of-use",
      notes="Franchised hotels (Comfort, Quality, Radisson Americas); watch RevPAR, unit growth, Radisson integration"),
    r("META", "Meta Platforms, Inc.",
      ir="https://investor.atmeta.com", newsroom="https://about.fb.com/news/",
      management_team="https://about.meta.com/company-info/",
      careers="https://www.metacareers.com", terms="https://www.facebook.com/legal/terms",
      notes="Watch Reels monetization, Threads growth, Reality Labs spend, AI capex, WhatsApp Business"),
    r("GOOG", "Alphabet Inc.",
      ir="https://abc.xyz/investor/", newsroom="https://blog.google",
      management_team="https://abc.xyz/investor/other/management/",
      careers="https://careers.google.com", terms="https://policies.google.com/terms",
      notes="Watch Search AI overviews impact, Cloud growth vs AWS/Azure, YouTube ad revenue, Waymo scaling, DOJ antitrust remedies"),
    r("SNAP", "Snap Inc.", "NYSE",
      ir="https://investor.snap.com", newsroom="https://newsroom.snap.com",
      management_team="https://investor.snap.com/corporate-governance/leadership/",
      careers="https://careers.snap.com", terms="https://snap.com/en-US/terms",
      notes="Watch DAU trend, Snapchat+ subscription, ad revenue vs guidance, AR partnerships"),
    r("PINS", "Pinterest, Inc.", "NYSE",
      ir="https://investor.pinterestinc.com", newsroom="https://newsroom.pinterest.com",
      careers="https://www.pinterestcareers.com", terms="https://policy.pinterest.com/en/terms-of-service",
      notes="Watch MAU trend, ARPU growth, shopping/commerce integrations, Amazon partnership scaling"),
    r("RDDT", "Reddit, Inc.", "NYSE",
      ir="https://investor.redditinc.com", newsroom="https://redditinc.com/press",
      careers="https://redditinc.com/careers", terms="https://www.redditinc.com/policies/user-agreement",
      reddit="https://www.reddit.com/r/reddit/",
      notes="Watch data licensing deals (Google/OpenAI), ad revenue growth, DAUq trend, Reddit Answers rollout"),
    r("APP",  "AppLovin Corporation",
      ir="https://investors.applovin.com", newsroom="https://www.applovin.com/blog/",
      careers="https://www.applovin.com/careers/", terms="https://www.applovin.com/legal/",
      notes="Ad tech / mobile gaming; watch AXON 2 ML model performance, e-commerce ad expansion, gaming vs non-gaming ad mix"),
    r("U",    "Unity Software Inc.", "NYSE",
      ir="https://investors.unity.com", newsroom="https://unity.com/company/press",
      careers="https://careers.unity.com", terms="https://unity.com/legal/terms-of-service",
      notes="Watch Runtime Fee resolution/aftermath, Create/Grow segment split, subscription renewals"),
    r("NFLX", "Netflix, Inc.",
      ir="https://ir.netflix.net", newsroom="https://about.netflix.com/en/news",
      careers="https://jobs.netflix.com", pricing="https://help.netflix.com/en/node/24926",
      terms="https://help.netflix.com/legal/termsofuse",
      notes="Watch subscriber adds, ARM changes, ads-tier growth, content spend, live sports strategy, password-sharing crackdown effect"),
    r("SPOT", "Spotify Technology S.A.", "NYSE",
      ir="https://investors.spotify.com", newsroom="https://newsroom.spotify.com",
      careers="https://www.lifeatspotify.com", pricing="https://www.spotify.com/premium",
      terms="https://www.spotify.com/legal/end-user-agreement/",
      notes="Watch MAU/Premium sub growth, gross margin (audiobook/podcast), pricing power, Marketplace ad revenue"),
    r("NYT",  "The New York Times Company", "NYSE",
      ir="https://investors.nytco.com", newsroom="https://www.nytco.com/press/",
      careers="https://www.nytco.com/careers/", terms="https://help.nytimes.com/hc/en-us/articles/115014893428",
      notes="Watch digital subs adds (news, cooking, games, Wirecutter, Athletic), ARPU, bundle attach rate"),
    r("LIF",  "LIF (verify ticker)", "?",
      notes="Ticker unclear — could be Life Insurance / other. Verify via SEC index."),
    r("BBY",  "Best Buy Co., Inc.", "NYSE",
      ir="https://investors.bestbuy.com", newsroom="https://corporate.bestbuy.com/newsroom/",
      careers="https://jobs.bestbuy.com", terms="https://www.bestbuy.com/site/help-topics/terms-and-conditions/pcmcat204400050067.c",
      notes="Watch comparable sales, membership (My Best Buy Total), TotalTech services attach, gaming/PC upgrade cycle"),
    r("DIS",  "The Walt Disney Company", "NYSE",
      ir="https://thewaltdisneycompany.com/investor-relations/",
      newsroom="https://thewaltdisneycompany.com/news/",
      careers="https://jobs.disneycareers.com",
      pricing="https://www.disneyplus.com/welcome",
      terms="https://disneytermsofuse.com",
      notes="Watch DTC (Disney+/Hulu/ESPN+) sub growth + profitability, ESPN direct-to-consumer 2025 launch, Parks OI, streaming ARPU"),
    r("FOXA", "Fox Corporation",
      ir="https://investor.foxcorporation.com", newsroom="https://www.foxcorporation.com/news/",
      careers="https://foxcareers.com", terms="https://www.fox.com/terms-of-use",
      notes="Watch Tubi ad growth, Fox News ratings + political cycle, Sports rights renewals, retransmission fees"),
    r("UMG",  "Universal Music Group N.V.", "Euronext Amsterdam",
      ir="https://www.universalmusic.com/investor-relations/", newsroom="https://www.universalmusic.com/news/",
      careers="https://www.universalmusic.com/careers/",
      notes="Non-US listed (Euronext); watch subscription streaming ARPU trends, TikTok deal, artist advance write-downs"),
    r("WMG",  "Warner Music Group Corp.",
      ir="https://investors.wmg.com", newsroom="https://www.wmg.com/news/",
      careers="https://careers.wmg.com", terms="https://www.wmg.com/terms-of-use",
      notes="Watch streaming revenue growth, mechanical royalty rates, artist advances, Live Nation vs. WMG dynamics"),
    r("LYV",  "Live Nation Entertainment, Inc.", "NYSE",
      ir="https://investors.livenationentertainment.com",
      newsroom="https://www.livenationentertainment.com/newsroom/",
      careers="https://careers.livenationentertainment.com",
      terms="https://help.ticketmaster.com/hc/en-us/articles/10736907984657",
      notes="Watch DOJ lawsuit outcomes, ticket AOV, concert attendance, Ticketmaster market share, VIP/premium seat mix"),
    r("STUB", "StubHub Holdings, Inc.", "NYSE",
      ir="https://investors.stubhub.com", newsroom="https://www.stubhub.com/press",
      careers="https://www.stubhub.com/careers/", terms="https://www.stubhub.com/legal/",
      notes="Recent 2024/2025 IPO; watch GMV growth, take rate, marketing spend, international expansion"),
    r("LION", "LION (verify ticker)", "?",
      notes="Ticker unclear — could be Fidelity National Financial or Lions Gate. Verify via SEC index."),
    r("OUT",  "OUTFRONT Media Inc.", "NYSE",
      ir="https://investor.outfrontmedia.com", newsroom="https://www.outfrontmedia.com/newsroom/",
      careers="https://www.outfrontmedia.com/careers", terms="https://www.outfrontmedia.com/legal/",
      notes="Out-of-home advertising (billboards, transit); watch same-store growth, digital conversion, MTA contract"),
    r("LAMR", "Lamar Advertising Company",
      ir="https://www.lamar.com/investor-relations", newsroom="https://www.lamar.com/news",
      careers="https://www.lamar.com/careers", terms="https://www.lamar.com/terms-of-use",
      notes="Out-of-home advertising REIT; watch same-store growth, digital board conversion, dividend policy"),
    r("TTWO", "Take-Two Interactive Software, Inc.",
      ir="https://ir.take2games.com", newsroom="https://ir.take2games.com/press-releases",
      careers="https://www.take2games.com/careers",
      notes="Watch GTA VI launch timing/reception, NBA 2K + Zynga performance, deferred revenue accretion"),
    r("EA",   "Electronic Arts Inc.",
      ir="https://ir.ea.com", newsroom="https://www.ea.com/news",
      careers="https://ea.com/careers", terms="https://tos.ea.com",
      notes="Watch EA SPORTS FC engagement, Battlefield launches, live services bookings, mobile gaming"),
    r("RBLX", "Roblox Corporation", "NYSE",
      ir="https://ir.roblox.com", newsroom="https://ir.roblox.com/news-releases",
      careers="https://careers.roblox.com", terms="https://en.help.roblox.com/hc/en-us/articles/115004647846",
      notes="Watch DAU trend (esp older cohorts), bookings/DAU, developer earnings share, advertising rollout, Roblox Studio AI"),
    r("INTU", "Intuit Inc.",
      ir="https://investors.intuit.com", newsroom="https://www.intuit.com/company/press-room/",
      careers="https://www.intuit.com/careers/", pricing="https://quickbooks.intuit.com/pricing/",
      terms="https://www.intuit.com/legal/terms-of-service/",
      notes="Watch QuickBooks Online growth, TurboTax mix, Credit Karma monetization, Mailchimp cross-sell, AI features"),
    r("ADP",  "Automatic Data Processing, Inc.",
      ir="https://investors.adp.com", newsroom="https://mediacenter.adp.com",
      careers="https://jobs.adp.com", terms="https://www.adp.com/about-adp/privacy.aspx",
      notes="Watch pays per control (employment proxy), client fund float, employer services growth, PEO segment"),
    r("PAYX", "Paychex, Inc.",
      ir="https://investor.paychex.com", newsroom="https://www.paychex.com/newsroom",
      careers="https://www.paychex.com/careers", terms="https://www.paychex.com/terms-of-use",
      notes="Watch client count, HR services attach rates, retention, float income, PEO expansion"),
    r("PAYC", "Paycom Software, Inc.", "NYSE",
      ir="https://investors.paycom.com", newsroom="https://www.paycom.com/resources/news/",
      careers="https://paycom.com/careers", terms="https://www.paycom.com/terms-of-use/",
      notes="Watch Beti self-service adoption, revenue per client, sales productivity, competitive pressure vs Rippling/Gusto"),
    r("PCTY", "Paylocity Holding Corporation",
      ir="https://investors.paylocity.com", newsroom="https://www.paylocity.com/resources/press-releases/",
      careers="https://www.paylocity.com/careers", terms="https://www.paylocity.com/terms/",
      notes="Watch mid-market win rate, revenue per client, community features / employee experience adoption"),
    r("CDW",  "CDW Corporation",
      ir="https://investor.cdw.com", newsroom="https://www.cdw.com/content/cdw/en/about/newsroom.html",
      careers="https://www.cdw.com/content/cdw/en/careers.html", terms="https://www.cdw.com/content/cdw/en/terms-conditions.html",
      notes="IT reseller; watch corporate/public/small-biz mix, gross margin, hardware refresh cycle, cloud services growth"),
]


def main() -> int:
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["URL Registry"]
    header = [c.value for c in ws[1]]
    col = {h: i + 1 for i, h in enumerate(header) if h}
    write_map = {
        "Ticker": "ticker", "Company": "company", "Exchange": "exchange",
        "Investor Relations": "ir", "Newsroom / Press": "newsroom",
        "Management Team": "management_team", "Careers / Jobs": "careers",
        "Product / Pricing": "pricing", "Terms & Conditions": "terms",
        "Regulatory Filings": "reg", "Reddit Community": "reddit",
        "Monitoring Notes": "notes",
    }
    existing = {
        str(ws.cell(row=r, column=col["Ticker"]).value or "").strip()
        for r in range(2, ws.max_row + 1)
    }
    added = 0
    for row in ROWS:
        if row["ticker"] in existing:
            print(f"  {row['ticker']:6s} already in xlsx — skip")
            continue
        new_row = ws.max_row + 1
        for hdr, key in write_map.items():
            if hdr in col:
                val = row.get(key, "")
                cell = ws.cell(row=new_row, column=col[hdr], value=val)
                if val and val.startswith("http"):
                    cell.hyperlink = Hyperlink(ref=cell.coordinate, target=val)
        if "Priority Tier" in col:
            ws.cell(row=new_row, column=col["Priority Tier"], value="v4")
        if "Active" in col:
            ws.cell(row=new_row, column=col["Active"], value=True)
        print(f"  {row['ticker']:6s} added at row {new_row}")
        added += 1
    wb.save(XLSX)
    print(f"\nAdded {added} v4 names.")
    return 0


if __name__ == "__main__":
    import sys
    sys.exit(main())
