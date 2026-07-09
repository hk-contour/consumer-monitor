"""Add the v3 coverage set — data analytics, auto retail, HSA custodian,
food services — to the URL Registry tab. Priority Tier = 'v3' so they
don't enter high/medium/all filters; use --ticker to scrape explicitly.

URL guesses use common patterns; some will 404 on first run. EDGAR
works immediately via SEC's ticker index for all 4.
"""

from pathlib import Path
import openpyxl
from openpyxl.worksheet.hyperlink import Hyperlink

XLSX = Path(__file__).resolve().parent.parent / "config" / "company_urls_jc.xlsx"

ROWS = [
    {
        "ticker": "VRSK", "company": "Verisk Analytics, Inc.", "exchange": "NASDAQ",
        "ir": "https://investor.verisk.com",
        "newsroom": "https://www.verisk.com/newsroom/",
        "management_team": "https://investor.verisk.com/corporate-governance/leadership",
        "careers": "https://careers.verisk.com",
        "pricing": "",
        "terms": "https://www.verisk.com/legal/terms-of-use/",
        "reg": "https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK=VRSK",
        "reddit": "",
        "notes": "Insurance analytics + data; watch subscription revenue growth, environmental/climate data expansion, insurance underwriting product launches",
    },
    {
        "ticker": "CVNA", "company": "Carvana Co.", "exchange": "NYSE",
        "ir": "https://investors.carvana.com",
        "newsroom": "https://www.carvana.com/blog",
        "management_team": "https://investors.carvana.com/corporate-governance/management",
        "careers": "https://www.carvana.com/careers",
        "pricing": "",
        "terms": "https://www.carvana.com/terms",
        "reg": "https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK=CVNA",
        "reddit": "https://www.reddit.com/r/carvana/",
        "notes": "Online used-car retailer; watch GPU (gross profit per unit), ADESA integration, retail unit growth vs guidance, auto finance spread",
    },
    {
        "ticker": "HQY", "company": "HealthEquity, Inc.", "exchange": "NASDAQ",
        "ir": "https://ir.healthequity.com",
        "newsroom": "https://ir.healthequity.com/news-events/press-releases",
        "management_team": "https://ir.healthequity.com/corporate-governance/leadership",
        "careers": "https://careers.healthequity.com",
        "pricing": "",
        "terms": "https://www.healthequity.com/terms/",
        "reg": "https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK=HQY",
        "reddit": "",
        "notes": "HSA custodian; watch HSA account growth, custodial yield spread, investment AUM balance, HSA-related legislation",
    },
    {
        "ticker": "ARMK", "company": "Aramark", "exchange": "NYSE",
        "ir": "https://www.aramark.com/about-us/investors",
        "newsroom": "https://www.aramark.com/about-us/news",
        "management_team": "https://www.aramark.com/about-us/leadership",
        "careers": "https://careers.aramark.com",
        "pricing": "",
        "terms": "https://www.aramark.com/legal/terms-of-use",
        "reg": "https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK=ARMK",
        "reddit": "",
        "notes": "Food services + uniforms; watch food inflation pass-through, uniform rental growth, education/business/healthcare segment mix",
    },
]


def main() -> int:
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["URL Registry"]
    header = [c.value for c in ws[1]]
    col = {h: i + 1 for i, h in enumerate(header) if h}

    write_map = {
        "Ticker": "ticker", "Company": "company", "Exchange": "exchange",
        "Investor Relations": "ir", "Newsroom / Press": "newsroom",
        "Management Team": "management_team", "Careers / Jobs": "careers",
        "Product / Pricing": "pricing", "Terms & Conditions": "terms",
        "Regulatory Filings": "reg", "Reddit Community": "reddit",
        "Monitoring Notes": "notes",
    }

    existing = {
        str(ws.cell(row=r, column=col["Ticker"]).value or "").strip()
        for r in range(2, ws.max_row + 1)
    }

    added = 0
    for row in ROWS:
        if row["ticker"] in existing:
            print(f"  {row['ticker']:6s} already in xlsx — skip")
            continue
        new_row = ws.max_row + 1
        for hdr, key in write_map.items():
            if hdr in col:
                val = row.get(key, "")
                cell = ws.cell(row=new_row, column=col[hdr], value=val)
                if val and val.startswith("http"):
                    cell.hyperlink = Hyperlink(ref=cell.coordinate, target=val)
        if "Priority Tier" in col:
            ws.cell(row=new_row, column=col["Priority Tier"], value="v3")
        if "Active" in col:
            ws.cell(row=new_row, column=col["Active"], value=True)
        print(f"  {row['ticker']:6s} added at row {new_row}")
        added += 1

    wb.save(XLSX)
    print(f"\nAdded {added} new names with tier='v3', Active=TRUE.")
    print("Trigger with: gh workflow run monitor.yml -f tickers=VRSK,CVNA,HQY,ARMK")
    return 0


if __name__ == "__main__":
    import sys
    sys.exit(main())
