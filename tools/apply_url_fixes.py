"""Apply Phase 0 URL fixes to the master xlsx.

Updates both cell.value and cell.hyperlink.target so the spreadsheet stays
clickable for the analyst.

Fixes sourced from the URL research agent's findings (2026-05-20). HIGH/MEDIUM
confidence applied; LOW confidence noted in comments below the FIXES dict.
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.worksheet.hyperlink import Hyperlink

XLSX = Path(__file__).resolve().parent.parent / "config" / "company_urls_jc.xlsx"

# (ticker, xlsx_column_header) -> new_url
FIXES: dict[tuple[str, str], str] = {
    ("SHOP",     "Management Team"):       "https://www.shopify.com/investors/board-of-directors",
    ("WIX",      "Newsroom / Press"):      "https://www.wix.com/press-room/home",
    ("WIX",      "Management Team"):       "https://investors.wix.com/board-of-directors",
    ("GLBE",     "Newsroom / Press"):      "https://www.global-e.com/media-centre/",
    ("GLBE",     "Product / Pricing"):     "https://www.global-e.com/platform/",
    ("GLBE",     "Terms & Conditions"):    "https://www.global-e.com/terms/",
    ("GDDY",     "Management Team"):       "https://aboutus.godaddy.net/about-us/team/default.aspx",
    ("EBAY",     "Management Team"):       "https://investors.ebayinc.com/corporate-governance/board-of-directors/default.aspx",
    ("EBAY",     "Product / Pricing"):     "https://www.ebay.com/sellercenter/selling/start-selling-on-ebay/seller-fees",
    # EBAY terms: URL is fine, TooManyRedirects was transient — leave as is.
    ("W",        "Management Team"):       "https://www.aboutwayfair.com/leadership",
    ("ETSY",     "Management Team"):       "https://investors.etsy.com/company-information/executive-team",
    ("CHWY",     "Newsroom / Press"):      "https://investor.chewy.com/news-and-events/news/default.aspx",
    ("CHWY",     "Management Team"):       "https://investor.chewy.com/governance/executive-management/default.aspx",
    # CHWY terms: keep as-is, agent said URL is live (transient timeout in audit).
    ("DASH",     "Management Team"):       "https://ir.doordash.com/governance/management/default.aspx",
    ("DASH",     "Product / Pricing"):     "https://merchants.doordash.com/en-us/products/merchant-pricing",
    ("CART",     "Newsroom / Press"):      "https://www.instacart.com/company/newsroom",
    ("CART",     "Product / Pricing"):     "https://www.instacart.com/instacart-plus",
    ("MTCH",     "Newsroom / Press"):      "https://ir.mtch.com/investor-relations/news-events/news-events/default.aspx",
    ("Z",        "Management Team"):       "https://www.zillowgroup.com/about-us/our-leaders/",
    ("Z",        "Terms & Conditions"):    "https://www.zillow.com/z/corp/terms/",  # LOW confidence — verify
    ("RKT",      "Management Team"):       "https://www.rocketcompanies.com/our-team/leadership/",
    # RKT terms: keep as-is, URL likely canonical but bot-protected.
    ("COMP",     "Management Team"):       "https://investors.compass.com/governance/management/default.aspx",
    ("COMP",     "Terms & Conditions"):    "https://www.compass.com/legal/terms-of-service/",
    ("APPF",     "Careers / Jobs"):        "https://www.appfolio.com/company/careers",
    ("APPF",     "Product / Pricing"):     "https://www.appfolio.com/pricing",
    ("XYZ",      "Management Team"):       "https://investors.block.xyz/governance/leadership/default.aspx",
    ("XYZ",      "Product / Pricing"):     "https://squareup.com/us/en/payments/our-fees",
    ("BILL",     "Newsroom / Press"):      "https://www.bill.com/press-release",
    ("BILL",     "Management Team"):       "https://www.bill.com/leadership",
    ("PYPL",     "Management Team"):       "https://about.pypl.com/who-we-are/executive-leadership/default.aspx",
    ("PYPL",     "Terms & Conditions"):    "https://www.paypal.com/us/legalhub/paypal/useragreement-full",
    ("TOST",     "Management Team"):       "https://pos.toasttab.com/leadership",
    ("ADYEN NA", "Newsroom / Press"):      "https://www.adyen.com/press-and-media",
    ("ADYEN NA", "Management Team"):       "https://www.adyen.com/about/team",
    ("ADYEN NA", "Regulatory Filings"):    "https://investors.adyen.com/financials",
    ("DAVE",     "Terms & Conditions"):    "https://dave.com/terms/",
    ("DAVE",     "Product / Pricing"):     "https://dave.com/extra-cash-account",
    ("ZIP AU",   "Newsroom / Press"):      "https://zip.co/investors/news",
    ("ZIP AU",   "Terms & Conditions"):    "https://zip.co/au/page/important-information",
    ("ZIP AU",   "Regulatory Filings"):    "https://zip.co/investors/asx-announcements",
    ("SEZL",     "Product / Pricing"):     "https://sezzle.com/premium/",
    ("KLAR",     "Management Team"):       "https://investors.klarna.com/governance/leadership/default.aspx",
    ("KLAR",     "Terms & Conditions"):    "https://www.klarna.com/us/terms-of-use/",
    ("HOOD",     "Product / Pricing"):     "https://robinhood.com/us/en/support/articles/trading-fees-on-robinhood/",
    ("HOOD",     "Terms & Conditions"):    "https://cdn.robinhood.com/assets/robinhood/legal/Robinhood-Customer-Agreement.pdf",
    ("UPST",     "Terms & Conditions"):    "https://www.upstart.com/terms",
    ("RH",       "Management Team"):       "https://ir.rh.com/corporate-governance/leadership",
    ("WSM",      "Terms & Conditions"):    "https://www.williams-sonoma.com/m/customer-service/terms.html",
    ("SGI",      "Product / Pricing"):     "https://www.tempurpedic.com/shop-mattresses/",
    # SGI terms: PDF on Q4 CDN — leave for now, would need PDF extraction.
    ("ENVA",     "Newsroom / Press"):      "https://ir.enova.com/news-events/news-releases/default.aspx",
    ("ENVA",     "Management Team"):       "https://ir.enova.com/corporate-governance/leadership-team/default.aspx",
    ("ENVA",     "Product / Pricing"):     "https://www.enova.com/brands/",
    ("ENVA",     "Terms & Conditions"):    "https://www.enova.com/terms-conditions/",
    ("ONON",     "Management Team"):       "https://investors.on-running.com/governance/default.aspx",
}

# URLs explicitly LEFT UNCHANGED with reason — these are scraper-side issues
# (bot blocking, transient timeouts, PDFs). Documented for future audit reruns.
LEFT_UNCHANGED = {
    ("EBAY", "Terms & Conditions"):   "URL is correct; TooManyRedirects was transient HEAD-method issue.",
    ("CHWY", "Terms & Conditions"):   "URL is correct; transient timeout in audit.",
    ("CSGP", "Terms & Conditions"):   "URL likely canonical; 403 from bot protection (Phase 1 problem).",
    ("RKT",  "Terms & Conditions"):   "URL likely canonical; 503 from bot protection (Phase 1 problem).",
    ("ZIP AU", "Management Team"):    "URL likely canonical; 530 Cloudflare block (Phase 1 problem).",
    ("SGI",  "Terms & Conditions"):   "Somnigroup terms are a PDF on Q4 CDN — needs PDF extraction support.",
}


def main() -> int:
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["URL Registry"]

    # Build header -> column index from row 1
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_idx = {h: i + 1 for i, h in enumerate(header_row) if h}
    if "Ticker" not in col_idx:
        print("ERROR: Ticker column not found", file=sys.stderr)
        return 1

    # Build ticker -> row index
    ticker_rows: dict[str, int] = {}
    for row_idx in range(2, ws.max_row + 1):
        tkr = ws.cell(row=row_idx, column=col_idx["Ticker"]).value
        if tkr:
            ticker_rows[str(tkr).strip()] = row_idx

    applied = 0
    skipped = []
    for (ticker, col_header), new_url in FIXES.items():
        if ticker not in ticker_rows:
            skipped.append(f"{ticker}/{col_header}: ticker not in xlsx")
            continue
        if col_header not in col_idx:
            skipped.append(f"{ticker}/{col_header}: column not in xlsx")
            continue
        cell = ws.cell(row=ticker_rows[ticker], column=col_idx[col_header])
        old_val = cell.value
        cell.value = new_url
        # Update or create hyperlink so the cell stays clickable
        if cell.hyperlink is not None:
            cell.hyperlink.target = new_url
        else:
            cell.hyperlink = Hyperlink(ref=cell.coordinate, target=new_url)
        applied += 1
        print(f"  {ticker:10s} {col_header:25s} OLD: {str(old_val)[:50]}")
        print(f"  {' ':10s} {' ':25s} NEW: {new_url}")

    wb.save(XLSX)
    print(f"\nApplied {applied} URL fixes")
    if skipped:
        print(f"Skipped {len(skipped)}:")
        for s in skipped:
            print(f"  {s}")
    if LEFT_UNCHANGED:
        print(f"\nLEFT UNCHANGED (scraper-side issues — see comments):")
        for (tkr, col), reason in LEFT_UNCHANGED.items():
            print(f"  {tkr:10s} {col:25s} {reason}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
