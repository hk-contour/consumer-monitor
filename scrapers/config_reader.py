"""Read the master xlsx URL Registry and return one structured record per company.

The xlsx is the single source of truth — no parallel JSON config. Analyst edits the
spreadsheet; the scraper reads it.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterator

import openpyxl

from .tiers import tier_for

DEFAULT_XLSX = Path(__file__).resolve().parent.parent / "config" / "company_urls_jc.xlsx"

URL_COLUMNS = [
    "investor_relations",
    "newsroom",
    "management_team",
    "careers",
    "pricing",
    "terms",
    "regulatory_filings",
    "reddit",
]

# (xlsx header -> our field name)
HEADER_MAP = {
    "Ticker": "ticker",
    "Company": "company",
    "Exchange": "exchange",
    "Investor Relations": "investor_relations",
    "Newsroom / Press": "newsroom",
    "Management Team": "management_team",
    "Careers / Jobs": "careers",
    "Product / Pricing": "pricing",
    "Terms & Conditions": "terms",
    "Regulatory Filings": "regulatory_filings",
    "Reddit Community": "reddit",
    "Monitoring Notes": "notes",
    "Newsroom RSS": "newsroom_rss",  # Phase 1C: confirmed RSS feed for newsroom
}

# Selector columns: column header -> source field name the selector applies to.
# Selectors narrow what static_pages.fetch_and_extract scrapes (Phase 1D).
SELECTOR_COLUMN_MAP = {
    "Pricing Selector": "pricing",
}


@dataclass
class Company:
    ticker: str
    company: str
    exchange: str
    tier: str  # "high" | "medium" | "blocked" | "unknown"
    active: bool = True
    urls: dict[str, str] = field(default_factory=dict)  # field_name -> url
    selectors: dict[str, str] = field(default_factory=dict)  # source -> CSS selector
    notes: str = ""
    analyst: str = ""  # coverage owner from the "Analyst" xlsx column ("1"/"2")

    def url(self, key: str) -> str | None:
        return self.urls.get(key)


def load_companies(xlsx_path: Path = DEFAULT_XLSX) -> list[Company]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    if "URL Registry" not in wb.sheetnames:
        raise ValueError(f"'URL Registry' tab not found in {xlsx_path}")

    ws = wb["URL Registry"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    header_idx = {h: i for i, h in enumerate(header) if h}

    out: list[Company] = []
    for row in rows:
        if not row or not row[header_idx.get("Ticker", 0)]:
            continue
        ticker = str(row[header_idx["Ticker"]]).strip()
        urls: dict[str, str] = {}
        for xlsx_header, field_name in HEADER_MAP.items():
            if field_name in {"ticker", "company", "exchange", "notes"}:
                continue
            idx = header_idx.get(xlsx_header)
            if idx is None:
                continue
            val = row[idx]
            if val and isinstance(val, str) and val.strip():
                urls[field_name] = val.strip()

        # Tier from xlsx column if present; fall back to tiers.py
        tier_val = (row[header_idx["Priority Tier"]]
                    if "Priority Tier" in header_idx else None)
        tier = str(tier_val).strip().lower() if tier_val else tier_for(ticker)

        # Active flag from xlsx; default True if column absent
        active_val = (row[header_idx["Active"]]
                      if "Active" in header_idx else True)
        active = active_val if active_val is not None else True

        # Coverage owner ("Analyst" column): "1" / "2"
        analyst_val = (row[header_idx["Analyst"]]
                       if "Analyst" in header_idx else None)
        analyst = str(analyst_val).strip() if analyst_val is not None else ""

        # Per-source CSS selectors (Phase 1D)
        selectors: dict[str, str] = {}
        for col_header, source_field in SELECTOR_COLUMN_MAP.items():
            idx = header_idx.get(col_header)
            if idx is None:
                continue
            val = row[idx]
            if val and isinstance(val, str) and val.strip():
                selectors[source_field] = val.strip()

        out.append(
            Company(
                ticker=ticker,
                company=str(row[header_idx.get("Company", 1)] or "").strip(),
                exchange=str(row[header_idx.get("Exchange", 2)] or "").strip(),
                tier=tier,
                active=bool(active),
                analyst=analyst,
                urls=urls,
                selectors=selectors,
                notes=str(row[header_idx["Monitoring Notes"]] or "").strip()
                if "Monitoring Notes" in header_idx
                else "",
            )
        )
    return out


def companies_by_tier(tier: str, xlsx_path: Path = DEFAULT_XLSX) -> Iterator[Company]:
    for c in load_companies(xlsx_path):
        if c.tier == tier:
            yield c


if __name__ == "__main__":
    cos = load_companies()
    print(f"Loaded {len(cos)} companies")
    by_tier: dict[str, int] = {}
    for c in cos:
        by_tier[c.tier] = by_tier.get(c.tier, 0) + 1
    print("By tier:", by_tier)
    for c in cos[:3]:
        print(f"  {c.ticker:6s} {c.tier:10s} {c.company} ({len(c.urls)} urls)")
